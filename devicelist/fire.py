from django.shortcuts import render
from django.conf import settings
from django.shortcuts import redirect
from django.core.mail import send_mail
from django.core.mail import mail_admins
from django.http import JsonResponse, FileResponse
from openpyxl import load_workbook

import sys, os, re, time, io
import datetime
import json
import mysql.connector
import fire_fias
try:
	from acds import configuration
except:
	import configuration

def db_model_search(query):
    try:
        conn = mysql.connector.connect(host=configuration.FIRE_HOST,
                                    database=configuration.FIRE_DB,
                                    user=configuration.FIRE_USER,
                                    password=configuration.FIRE_PASS)
    except Exception as e:
        print(e)
        return
    try:
        cursor = conn.cursor()
        cursor.execute(query)
        row = cursor.fetchall()
        return row
    except Exception as e:
        print(e)
    
    finally:
        cursor.close()
        conn.close()

def db_insert(query):
    try:
        conn = mysql.connector.connect(host=configuration.FIRE_HOST,
                                    database=configuration.FIRE_DB,
                                    user=configuration.FIRE_USER,
                                    password=configuration.FIRE_PASS)
    except Exception as e:
        return str(e)
    try:
        cursor = conn.cursor()
        cursor.execute(query)
        lastid = cursor.lastrowid
        conn.commit()
        return lastid
    except Exception as e:
        return str(e)
    
    finally:
        cursor.close()
        conn.close()

def firelist(request):
  if not request.user.is_authenticated:
    return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))
  if request.method == 'GET':
    fi_fi = request.GET.get("fire_fias", "")
    fi_se = request.GET.get("fire_serial", "")
    fi_in = request.GET.get("fire_inventory", "")
#fire_fias={"region_fias_id":"54049357-326d-4b8f-b224-3c6dc25d6dd3","area_fias_id":null,"city_fias_id":"9ae64229-9f7b-4149-b27a-d1f6ec74b5ce","settlement_fias_id":null,"street_fias_id":"b5701907-1537-4e52-b93a-d566a47086f7","house_fias_id":"76bedc4e-a5cb-42d3-bbea-8b2765bdd14c"}
    firedic={} # потом сделаю поиск по серийнику и инвентарному номеру, может быть
    if (fi_fi) :
        fii = json.loads(fi_fi)  # составляем sql запрос адреса географического. Если есть значение задаём его в явном виде в запрос
        if fii["region_fias_id"]:
           fi_region = "ff.region_fias_id LIKE '%"+fii["region_fias_id"]+"%'"
        else:
           fi_region = "ff.region_fias_id is NULL"           
        if fii["area_fias_id"]:
           fi_area = "ff.area_fias_id LIKE '%"+fii["area_fias_id"]+"%'"
        else:
           fi_area = "ff.area_fias_id is NULL"
        if fii["city_fias_id"] :
           fi_city = "ff.city_fias_id LIKE '%"+fii["city_fias_id"]+"%'"
        else:
           fi_city = "ff.city_fias_id is NULL"
        if fii["settlement_fias_id"]:
            fi_sett = "ff.settlement_fias_id LIKE '%"+fii["settlement_fias_id"] +"%'"
        else:
            fi_sett = "ff.settlement_fias_id is NULL"
        if fii["street_fias_id"]:
            fi_street = "ff.street_fias_id LIKE '%"+fii["street_fias_id"]+"%'"
        else:
            fi_street = "ff.street_fias_id is NULL"
        if fii["house_fias_id"]:
            fi_house = "ff.house_fias_id LIKE '%"+fii["house_fias_id"]+"%'"
        else:
            fi_house = "ff.house_fias_id is NULL"
           
        query_fire = f"""select fl.fireid, fl.type, fl.serial, fl.inventory, fl.room, fl.fullweight, fl.status, fl.comandor, fl.address, fl.ClassList 
        from FireSupressor.FireList as fl
			  LEFT JOIN FireSupressor.FireFias as ff on fl.fireid = ff.fireid
			  WHERE {fi_region}  and  {fi_area} and {fi_city} and {fi_sett} and {fi_street} and {fi_house}"""
        sqldata = db_model_search(query_fire)
        for iii in sqldata:
             querycheck =  "select fc.chargeid, fc.Chargedata, fc.Checkdata, fc.weight, fc.userwho from FireCheck as fc where fc.fireid = " + str(iii[0])
             checklist = db_model_search(querycheck)
             firedic[iii[0]]=({"fire":iii, "check":checklist})
        return render(request, 'wbs/fire.html', locals())
        
			

  firedic={}
  sqldata= db_model_search("select fireid, type, serial, inventory, room, fullweight, status, comandor, address, ClassList from FireSupressor.FireList")
  for iii in sqldata:
    querycheck =  "select fc.chargeid, fc.Chargedata, fc.Checkdata, fc.weight, fc.userwho from FireCheck as fc where fc.fireid = " + str(iii[0])
    checklist = db_model_search(querycheck)
    firedic[iii[0]]=({"fire":iii, "check":checklist})
  
  
  return render(request, 'wbs/fire.html', locals())


#self.sql = mysql.connector.connect(host = server, user = login_sql, password = p_sql, charset='utf8', use_unicode = True)
#			self.cursor_mysql = self.sql.cursor(buffered=True) 
  
  
  
  
def set_fire_data(request):
  if not (request.user.is_authenticated):
    return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))
  group = [ x.name for x in request.user.groups.all()]
  if not('admins' in group or 'engineers' in group):
    return JsonResponse({'error': 'You have not permission'}, safe=False)

  firedic = {"carbonic":"ОУ", "powdery":"ОП", "water":"ОВМ", "ready":"Готов", "backup":"на складе", "lost":"потерян", "empty":"разряжен", "broken":"сломан", "fireA":"1", "fireB":"2", "fireC":"3","fireD":"4", "fireE":"5"}
  if request.method == 'GET':
    all_data = json.loads(request.GET['all_data'])
    if all_data['action'] =="insertnew":
      address = all_data['address']
      room = all_data['room']
      comandor = all_data['comandor']
      serial = all_data['serial']
      inventory = all_data['inventory']
      weight = all_data['weight']
      firetype = firedic[all_data['firetype']]
      firestatus = firedic[all_data['firestatus']]
      fireclass = all_data['fireclass'].split(',')
      if weight =="":
        insertquery = "INSERT INTO FireSupressor.FireSupressor (serial, inventory, type,  room, comandor, status) VALUES ('"+serial+"', '"+inventory+"', '"+firetype+"', '"+room+"', '"+comandor+"', '"+firestatus+"')"
      else:
        insertquery = "INSERT INTO FireSupressor.FireSupressor (serial, inventory, type, fullweight, room, comandor, status) VALUES ('"+serial+"', '"+inventory+"', '"+firetype+"', '"+weight+"', '"+room+"', '"+comandor+"', '"+firestatus+"')"
      lastid_ee=db_insert(insertquery)
      fias_rez=fire_fias.fire_fias_insert(lastid_ee,address)

      
      for ttt in fireclass:
        insertclass = "INSERT INTO FireSupressor.FireClassAndFireID (fireid, classid) VALUES ('"+str(lastid_ee)+"', '"+firedic[ttt]+"')"
        print(insertclass)
        db_insert(insertclass)
      
      return JsonResponse({'insert': 'ok', "result":lastid_ee, "fias":fias_rez}, safe=False)
      

    #t.count_website(t, page='set_wbs_data', username=request.user.username)
  else:
    return JsonResponse({'error': 'wrong method'}, safe=False)

  if request.method == 'GET':
    all_data = json.loads(request.GET['all_data'])
    if all_data['action'] =="insertcheck":
      fireid = all_data['fireid']
      weight = all_data['weight']
      data_check = all_data['data']
      comandor = all_data['comandor']
      firecheck_type = all_data['firecheck_type']
      if weight =="":
        if firecheck_type == "chardged":
          insertcheck = "INSERT INTO FireSupressor.FireCheck (fireid, Chargedata, userwho) VALUES ('"+fireid+"', '"+data_check+"', '"+comandor+"')"
        else:
          insertcheck = "INSERT INTO FireSupressor.FireCheck (fireid, Checkdata, userwho) VALUES ('"+fireid+"', '"+data_check+"', '"+comandor+"')"
      else:
        if firecheck_type == "chardged":
          insertcheck = "INSERT INTO FireSupressor.FireCheck (fireid, Chargedata, userwho, weight) VALUES ('"+fireid+"', '"+data_check+"', '"+comandor+"', '"+weight+"')"
        else:
          insertcheck = "INSERT INTO FireSupressor.FireCheck (fireid, Checkdata, userwho, weight) VALUES ('"+fireid+"', '"+data_check+"', '"+comandor+"', '"+weight+"')"
      eecheck=db_insert(insertcheck)
      return JsonResponse({'insert': 'ok', "last_id":eecheck}, safe=False)

  if request.method == 'GET':
    all_data = json.loads(request.GET['all_data'])
    if all_data['action'] =="editfire":
      fireid = all_data['fireid']
      address = all_data['address']
      room = all_data['room']
      comandor = all_data['comandor']
      serial = all_data['serial']
      inventory = all_data['inventory']
      weight = all_data['weight']
      firetype = firedic[all_data['firetype']]
      firestatus = firedic[all_data['firestatus']]
      fireclass = all_data['fireclass'].split(',')
      updatequery = "UPDATE FireSupressor.FireSupressor SET serial='"+serial+"', inventory='"+inventory+"', type='"+firetype+"',  room='"+room+"', comandor='"+comandor+"', status='"+firestatus+"' WHERE fireid="+fireid+" "
      lastid_ee=db_insert(updatequery)
      fias_rez=fire_fias.fire_fias_update(fireid,address)  
      # дописать обновление данных о классах огнетушителя
      return JsonResponse({'edit': 'ok',"result":lastid_ee}, safe=False)    

  return JsonResponse({'update': 'unknown'}, safe=False)
  
def firejornal(request):
  if not request.user.is_authenticated:
    return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))
  if request.method == 'GET':
    fireliststr = request.GET.get("firelist", "")  #"dict_keys([6, 3, 9, 4, 10, 7, 12, 5, 14])"
    firelist1 = fireliststr.replace("dict_keys","")
    firelist2 = firelist1.replace("[","")
    firelist3 = firelist2.replace("]","")
    sqlselect1 = "select fireid, type, serial, inventory, room, fullweight, status, comandor, address, ClassList from FireSupressor.FireList WHERE fireid IN "+firelist3

  try:
    wb1 = load_workbook('static/devicelist/firejornal2.xlsx')
  except Exception as e:
    return JsonResponse({'error': str(e)}, safe=False)

  buffer = io.BytesIO()

  sheet1 = wb1['fire1']
  #firedic={}
  sqldata= db_model_search(sqlselect1)
  cellnn=0
  if sqldata:
    for iii in sqldata:
      querycheck =  "select fc.chargeid, fc.Chargedata, fc.Checkdata, fc.weight, fc.userwho from FireCheck as fc where fc.fireid = " + str(iii[0])
      checklist = db_model_search(querycheck)
      if iii[1]: 
            fire_type = str(iii[1])
      else:
            fire_type = ""
      if iii[2] :
            fire_serial = str(iii[2])
      else:
            fire_serial = ""
      if iii[3] :
            fire_inventory = str(iii[3])
      else:
            fire_inventory = ""
      if iii[6]:
         fire_status = str(iii[6])
      else:
         fire_status = ""
      if checklist:
        if kkk[1]:
          charge_data = str(kkk[1])
        else:
          charge_data =""
        if kkk[2]:
          check_data = str(kkk[2])
        else:
          check_data = ""
        if kkk[4]:
          user_who = str(kkk[4])
        else:
          user_who =""
        for kkk in checklist:
          sheet1.cell(column=1, row=5+cellnn, value=fire_serial+" "+fire_inventory + " тип:" + fire_type)
          sheet1.cell(column=7, row=5+cellnn, value=fire_status)
          sheet1.cell(column=5, row=5+cellnn, value=charge_data) # дата заправки
          sheet1.cell(column=2, row=5+cellnn, value=check_data) # дата проверки
          sheet1.cell(column=9, row=5+cellnn, value=user_who) # ответственный
          cellnn+=1
      else:
        sheet1.cell(column=1, row=5+cellnn, value=fire_serial+" "+fire_inventory + " тип:" + fire_type)
        sheet1.cell(column=7, row=5+cellnn, value=fire_status) # статус
        cellnn+=1
  wb1.save(buffer)
  buffer.seek(0)
  
  #return JsonResponse({'sqlselect': str(sqldata)}, safe=False)
  return FileResponse(buffer,as_attachment=True, filename='firejornal.xlsx')

  
  
  
  
  
  
  
  
  
  
  
  
